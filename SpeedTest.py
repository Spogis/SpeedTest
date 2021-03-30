# Python program to test 
# internet speed 
  
import speedtest  
import time
from openpyxl import load_workbook
from datetime import datetime

speedtester = speedtest.Speedtest()

SleepTime = 10 #Minutes 
TotalTime = 96 #Hours

NumberOfData = int(TotalTime*60/SleepTime)

# Start by opening the spreadsheet and selecting the main sheet
workbook = load_workbook(filename="SpeedTestData.xlsx")
sheet = workbook.active


def get_final_speed():
    speedtester.get_best_server()
    SpeedDownload = speedtester.download() / 1e+6
    SpeedDownload = round(SpeedDownload,2)
    
    SpeedUpload = speedtester.upload() / 1e+6
    SpeedUpload = round(SpeedUpload, 2)
    
    SpeedPing = round(speedtester.results.ping, 1)
    return [SpeedPing, SpeedDownload, SpeedUpload];

 
for i in range(NumberOfData):
    x=get_final_speed()
    rows = sheet.max_row
    print("Time %s SpeedTest %s" % (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), x))
    sheet.cell(row=rows+1, column=1).value = datetime.now()
    sheet.cell(row=rows+1, column=2).value =x[0]
    sheet.cell(row=rows+1, column=3).value =x[1]
    sheet.cell(row=rows+1, column=4).value =x[2]
    workbook.save(filename="SpeedTestData.xlsx")
    time.sleep(SleepTime*60)